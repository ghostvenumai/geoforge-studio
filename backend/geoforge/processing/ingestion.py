from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any, cast

import polars as pl
from charset_normalizer import from_bytes
from openpyxl import load_workbook


def detect_text_format(path: Path) -> tuple[str, str]:
    sample = path.read_bytes()[:128_000]
    try:
        sample.decode("utf-8")
        encoding = "utf-8-sig" if sample[:3] == bytes((0xEF, 0xBB, 0xBF)) else "utf-8"
    except UnicodeDecodeError:
        best = from_bytes(sample).best()
        candidate = best.encoding if best and best.encoding else "cp1252"
        normalized = candidate.casefold().replace("-", "_")
        western = ("cp125", "windows_125", "iso8859", "latin")
        encoding = candidate if normalized.startswith(western) else "cp1252"
    decoded = sample.decode(encoding, errors="replace")
    try:
        delimiter = csv.Sniffer().sniff(decoded[:16_384], delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    return encoding, delimiter


def _read_csv(path: Path) -> tuple[pl.DataFrame, str, str]:
    encoding, delimiter = detect_text_format(path)
    raw = path.read_bytes().decode(encoding, errors="replace").encode("utf-8")
    frame = pl.read_csv(
        io.BytesIO(raw),
        separator=delimiter,
        infer_schema_length=10_000,
        try_parse_dates=True,
        truncate_ragged_lines=False,
    )
    return frame, encoding, delimiter


def _read_xlsx(path: Path) -> pl.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("XLSX workbook has no active worksheet")
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError("XLSX worksheet is empty") from exc
    columns = [
        str(value) if value is not None else f"column_{index + 1}"
        for index, value in enumerate(header)
    ]
    records = [dict(zip(columns, row, strict=False)) for row in rows]
    workbook.close()
    return (
        pl.DataFrame(records, infer_schema_length=None) if records else pl.DataFrame(schema=columns)
    )


def read_dataset(path: Path, file_format: str) -> tuple[pl.DataFrame, str | None, str | None]:
    normalized = file_format.casefold()
    if normalized == "csv":
        return _read_csv(path)
    if normalized == "parquet":
        return pl.read_parquet(path), None, None
    if normalized in {"jsonl", "ndjson"}:
        return pl.read_ndjson(path), "utf-8", None
    if normalized == "json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            payload = payload.get("records", [payload])
        if not isinstance(payload, list):
            raise ValueError("JSON input must contain an array or a records array")
        return pl.DataFrame(payload, infer_schema_length=None), "utf-8", None
    if normalized == "xlsx":
        return _read_xlsx(path), None, None
    raise ValueError(f"Unsupported dataset format: {file_format}")


def schema_payload(frame: pl.DataFrame) -> dict[str, Any]:
    return {name: str(dtype) for name, dtype in frame.schema.items()}


def safe_preview(frame: pl.DataFrame, rows: int = 100) -> list[dict[str, Any]]:
    return cast(list[dict[str, Any]], json.loads(frame.head(rows).write_json()))
